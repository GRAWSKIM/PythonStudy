from urllib.request import Request, urlopen
from bs4 import BeautifulSoup
import openpyxl

url_list = list()


# 간략검색에서 얻은 데이터로 상세페이지 url을 만든다
def godetailproductnotagepopup(ejkgb, linkclass, barcode):
    return "http://www.kyobobook.co.kr/product/detailViewKor.laf?mallGb=KOR" + "&ejkGb=" + ejkgb + "&linkClass=" + \
           linkclass + "&barcode=" + barcode


def geturlfromtag(tag, detaillist):
    for a in tag:
        param = a.attrs['href']
        param = param.replace("javascript:goDetailProductNotAgePopup(", "")
        param = param.replace(")", "")
        param = param.replace("'", "")
        param = param.replace(" ", "")
        params = param.split(",")

        detaillist.append(godetailproductnotagepopup(params[0], params[1], params[2]))


# 문자열에서 데이터를 제외한 문자를 제거한다
def erasehtmltag(tag_str):
    tag_str = tag_str.replace("\t", "").replace("\r\n", "").replace("\r", "").replace("\n", "")

    while True:
        start_index = tag_str.find("<")
        end_index = tag_str.find(">")
        if start_index < 0 and end_index < 0:
            break

        replace_string = tag_str[start_index:end_index + 1]
        tag_str = tag_str.replace(replace_string, "")

    return tag_str


def get_file_name():
    return ''


bestseller_url = "http://www.kyobobook.co.kr/categoryRenewal/categoryMain.laf?perPage=50&mallGb=KOR&linkClass=01" \
                 "&menuCode=002 "
req = Request(bestseller_url)

res = urlopen(req)
html = res.read().decode('cp949')

bs = BeautifulSoup(html, 'html.parser')
tags = bs.findAll('a', attrs={'class': 'btn_new_window'})

geturlfromtag(tags, url_list)
# url 중복제거
unique_url_list = list(set(url_list))

# Create a excel file
wb = openpyxl.Workbook()
sheet = wb.worksheets[0]

# 열명
sheet.cell(row=1, column=1).value = "서명"
sheet.cell(row=1, column=2).value = "저자"
sheet.cell(row=1, column=3).value = "발행자"
sheet.cell(row=1, column=4).value = "발행년"
sheet.cell(row=1, column=5).value = "가격"
sheet.cell(row=1, column=6).value = "낱권ISBN"

# 데이터 파싱 -> excel 입력
for url in unique_url_list:
    req = Request(url)
    res = urlopen(req)
    html = res.read().decode('cp949')
    bs = BeautifulSoup(html, 'html.parser')
    # 총서명
    title_info_all = ""
    title = bs.find('h1', attrs={'class': 'title'})

    for factor in title:
        factor = erasehtmltag(str(factor))
        title_info_all = title_info_all + factor + " "

    title_info_all = title_info_all.strip()
    print(title_info_all)
    # 저자
    author = bs.find('span', attrs={'class': 'name'})
    author_info = author.text.replace("\t", "").replace("\r\n", "").replace("\r", "").replace("\n", "")
    print(author_info)
    # 발행자
    publisher = bs.find('span', attrs={'title': '출판사'})
    publisher_info = publisher.text.replace("\t", "").replace("\r\n", "").replace("\r", "").replace("\n", "")
    print(publisher_info)
    # 발행년
    publish_year = bs.find('span', attrs={'title': '출간일'})
    publish_year = publish_year.text.replace("\t", "").replace("\r\n", "").replace("\r", "").replace("\n", "")
    publish_year_info = publish_year[:4]
    print(publish_year_info)
    # 가격
    price = bs.find('span', attrs={'class': 'org_price'})
    price_info = price.text.replace("\t", "").replace("\r\n", "").replace("\r", "").replace("\n", "")\
        .replace("원", "").replace(",", "")
    print(price_info)
    # 낱권ISBN
    isbn = bs.find('span', attrs={'title': 'ISBN-13'})
    isbn_info = isbn.text.replace("\t", "").replace("\r\n", "").replace("\r", "").replace("\n", "")
    print(isbn_info)
    sheet.append([title_info_all, author_info, publisher_info, publish_year_info, price_info, isbn_info])

# 바탕화면에 저장
wb.save('C:\\Users\\pc\\Desktop\\bookinfo.xlsx')
"""
detail_html = BeautifulSoup(detail_page, 'html.parser')
#서명, 편/권차, 저자, 발행자, 발행년, 총서명, 총서편차, 책수, 가격. 낱권ISBN, 면장수, 물리적특성
"""
